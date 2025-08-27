def generer_grille(fichier, nom_feuille):
    import openpyxl

    excel = openpyxl.load_workbook(fichier, data_only=True)

    feuille = excel[nom_feuille]

    padding_de_grille = '|&#8288 {: style="padding:0"}'

    grille = []

    # Écrire les entêtes du tableau de l'horaire
    entete = ""
    trait = ""
    for col in range(1, feuille.max_column + 1):
        entete += feuille.cell(row=1, column=col).value
        trait += "--"
        if col < feuille.max_column:
            entete += "|"
            trait += "|"
        else:
            entete += "\n"
            trait += "\n"

    grille.append(entete)
    grille.append(trait)

    for row in range(2, feuille.max_row + 1):
        ligne = ""
        has_colspan = False
        for col in range(1, feuille.max_column + 1):
            valeur = feuille.cell(row=row, column=col).value
            if valeur is None:
                valeur = ""
            else:
                valeur = str(valeur)
            ligne += valeur
            if col < feuille.max_column:
                if not has_colspan:
                    ligne += "|"
            else:
                if has_colspan:
                    ligne += padding_de_grille * (feuille.max_column - 1)
                ligne += "\n"
            if "colspan" in valeur:
                has_colspan = True

        

        grille.append(ligne)
    
    return grille

def generer_projet_integrateur():

    grille = generer_grille("template/grille_correction.xlsx", "grille")

    with open("./wiki/projet-de-session.md", "w") as f:

        with open("template/projet-de-session.md", "r") as t:
            f.writelines(t.readlines())

        f.writelines(grille)

def generer_horaire_rencontre():

    grille = generer_grille("template/horaire-rencontre.xlsx", "grille")

    with open("./wiki/horaire-rencontre.md", "w") as f:
        f.writelines(grille)


print("Générer la page du projet de session")
generer_projet_integrateur()
print("Générer la page de l'horaire de rencontre")
generer_horaire_rencontre()